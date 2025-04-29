#!/usr/bin/env python3
"""
Transaction Reports Generator
Generates Báo Cáo Mua Vào (Incoming) and Báo Cáo Bán Ra (Outgoing) reports from processed transaction data
"""
import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import shutil
import logging


class TransactionReportGenerator:
    """
    Generates separate incoming and outgoing transaction reports
    based on the direction attribute in the processed data
    """

    def __init__(self, input_file="flattened_attachments.csv", template_file=None):
        """
        Initialize with input data file and optional template

        Args:
            input_file (str): Path to the CSV file containing transaction data
            template_file (str, optional): Path to Excel template for reports
        """
        self.input_file = input_file
        self.template_file = template_file
        self.data = None
        self.logger = self._setup_logger()

    def _setup_logger(self):
        """Setup basic logger"""
        logger = logging.getLogger("TransactionReports")
        if not logger.handlers:
            logger.setLevel(logging.INFO)
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        return logger

    def load_data(self):
        """Load and validate the transaction data"""
        try:
            self.logger.info(f"Loading data from {self.input_file}")
            self.data = pd.read_csv(self.input_file, encoding="utf-8-sig")

            # Check if direction column exists
            if "direction" not in self.data.columns:
                self.logger.error("Missing 'direction' column in the data")
                return False

            # Fill any missing directions with 'UNKNOWN' - fixed to avoid FutureWarning
            # Instead of using chained assignment with inplace=True
            self.data = (
                self.data.copy()
            )  # Create a copy to ensure we're modifying the original
            self.data["direction"] = self.data["direction"].fillna("UNKNOWN")

            # Basic validation
            self.logger.info(f"Loaded {len(self.data)} transactions")
            self.logger.info(
                f"Directions found: {self.data['direction'].value_counts().to_dict()}"
            )

            return True
        except Exception as e:
            self.logger.error(f"Error loading data: {str(e)}")
            return False

    def generate_incoming_report(self, output_file="Bao_Cao_Ban_Ra.xlsx"):
        """
        Generate the incoming transactions report (Báo Cáo Bán Ra)
        Args:
            output_file (str): Path to save the generated report
        Returns:
            bool: Success or failure
        """
        if self.data is None:
            if not self.load_data():
                return False
        try:
            # Filter for incoming transactions
            incoming_df = self.data[self.data["direction"] == "INCOMING"].copy()
            self.logger.info(f"Processing {len(incoming_df)} incoming transactions")
            if len(incoming_df) == 0:
                self.logger.warning("No incoming transactions found")
                return False
            # Sort by date
            if "date" in incoming_df.columns:
                incoming_df["date"] = pd.to_datetime(
                    incoming_df["date"], errors="coerce"
                )
                incoming_df.sort_values("date", inplace=True)
            # Create report
            return self._create_report(incoming_df, output_file, report_type="BÁN RA")
        except Exception as e:
            self.logger.error(f"Error generating incoming report: {str(e)}")
            return False

    def generate_outgoing_report(self, output_file="Bao_Cao_Mua_Vao.xlsx"):
        """
        Generate the outgoing transactions report (Báo Cáo Mua Vào)
        Args:
            output_file (str): Path to save the generated report
        Returns:
            bool: Success or failure
        """
        if self.data is None:
            if not self.load_data():
                return False
        try:
            # Filter for outgoing transactions
            outgoing_df = self.data[self.data["direction"] == "OUTGOING"].copy()
            self.logger.info(f"Processing {len(outgoing_df)} outgoing transactions")
            if len(outgoing_df) == 0:
                self.logger.warning("No outgoing transactions found")
                return False
            # Sort by date
            if "date" in outgoing_df.columns:
                outgoing_df["date"] = pd.to_datetime(
                    outgoing_df["date"], errors="coerce"
                )
                outgoing_df.sort_values("date", inplace=True)
            # Create report
            return self._create_report(outgoing_df, output_file, report_type="MUA VÀO")
        except Exception as e:
            self.logger.error(f"Error generating outgoing report: {str(e)}")
            return False

    def generate_all_reports(self, base_folder="reports"):
        """
        Generate both incoming and outgoing reports
        Args:
            base_folder (str): Base folder to save reports
        Returns:
            bool: Success or failure
        """
        try:
            if not os.path.exists(base_folder):
                os.makedirs(base_folder)
                self.logger.info(f"Created reports folder: {base_folder}")
            # Generate both reports
            outgoing_file = os.path.join(base_folder, "Bao_Cao_Mua_Vao.xlsx")
            incoming_file = os.path.join(base_folder, "Bao_Cao_Ban_Ra.xlsx")
            incoming_success = self.generate_incoming_report(incoming_file)
            outgoing_success = self.generate_outgoing_report(outgoing_file)
            if incoming_success or outgoing_success:
                self.logger.info("Report generation completed")
                return True
            else:
                self.logger.error("Failed to generate any reports")
                return False
        except Exception as e:
            self.logger.error(f"Error generating reports: {str(e)}")
            return False

    def _create_report(self, df, output_file, report_type):
        """
        Create an Excel report with proper formatting that exactly matches the BKMV template
        with Vietnamese text, center alignment, and proper borders
        Args:
            df: DataFrame containing the filtered transaction data
            output_file: Path to save the output Excel file
            report_type: "MUA VÀO" or "BÁN RA"
        Returns:
            bool: Success or failure
        """
        try:
            # Get entity name and tax number
            entity_name = "CÔNG TY TNHH"
            entity_tax_number = ""
            if "entity_name" in df.columns and len(df["entity_name"].dropna()) > 0:
                entity_names = df["entity_name"].dropna().unique()
                if len(entity_names) > 0:
                    entity_name = entity_names[0]

            if (
                "entity_tax_number" in df.columns
                and len(df["entity_tax_number"].dropna()) > 0
            ):
                tax_numbers = df["entity_tax_number"].dropna().unique()
                if len(tax_numbers) > 0:
                    entity_tax_number = str(tax_numbers[0])

            # Create new workbook
            self.logger.info("Creating new workbook with BKMV template format")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "BKMV"

            # Set column widths based on template
            column_widths = {
                "A": 8,  # STT
                "B": 15,  # Hóa đơn chứng từ mua - Seri
                "C": 10,  # Số HĐ
                "D": 10,  # Ngày HĐ
                "E": 25,  # Tên người bán
                "F": 15,  # Mã số thuế
                "G": 20,  # Mặt hàng
                "H": 15,  # Doanh số mua chưa thuế
                "I": 8,  # Thuế suất
                "J": 15,  # Thuế GTGT
                "K": 15,  # Ghi chú
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # ROW 1: Main title
            title = (
                "BẢNG KÊ HOÁ ĐƠN CHỨNG TỪ HÀNG HOÁ, DỊCH VỤ MUA VÀO"
                if report_type == "MUA VÀO"
                else "BẢNG KÊ HOÁ ĐƠN CHỨNG TỪ HÀNG HOÁ, DỊCH VỤ BÁN RA"
            )
            sheet.merge_cells("A1:K1")
            cell = sheet.cell(row=1, column=1)
            cell.value = title
            cell.font = Font(name="Times New Roman", size=14, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 25

            # ROW 2: Subtitle 1
            sheet.merge_cells("A2:K2")
            cell = sheet.cell(row=2, column=1)
            cell.value = "Kèm theo tờ khai thuế GTGT"
            cell.font = Font(name="Times New Roman", size=12, bold=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[2].height = 20

            # ROW 3: Subtitle 2
            sheet.merge_cells("A3:K3")
            cell = sheet.cell(row=3, column=1)
            cell.value = "(Dùng cho cơ sở kê khai khấu trừ thuế hàng tháng)"
            cell.font = Font(name="Times New Roman", size=12, bold=False, italic=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[3].height = 20

            # ROW 4: Period
            current_year = datetime.now().year
            current_quarter = (datetime.now().month - 1) // 3 + 1

            sheet.merge_cells("A4:K4")
            cell = sheet.cell(row=4, column=1)
            cell.value = f"Kỳ quý Năm {current_year}"
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[4].height = 20

            # ROW 5: Company name and tax code
            # Left side - Company name (with text overflow)
            sheet.merge_cells("A5:G5")
            cell = sheet.cell(row=5, column=1)
            cell.value = f"Tên cơ sở kinh doanh : {entity_name}"
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=False
            )

            # Right side - Tax code (moved further right)
            sheet.merge_cells("H5:K5")
            cell = sheet.cell(row=5, column=8)
            cell.value = f"Mã số thuế : {entity_tax_number}"
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[5].height = 20

            # ROW 6: Address (optional)
            sheet.merge_cells("A6:K6")
            cell = sheet.cell(row=6, column=1)
            cell.value = "Địa chỉ : "
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[6].height = 20

            # ROW 7: Table headers
            headers_row1 = {
                "A7": "STT",
                "B7": "Hóa đơn chứng từ mua",
                "E7": "Tên người bán" if report_type == "MUA VÀO" else "Tên người mua",
                "F7": "Mã số thuế",
                "G7": "Mặt hàng",
                "H7": (
                    "Doanh số mua\nchưa thuế"
                    if report_type == "MUA VÀO"
                    else "Doanh số bán\nchưa thuế"
                ),
                "I7": "Thuế\nsuất",
                "J7": "Thuế\nGTGT",
                "K7": "Ghi\nchú",
            }

            # Merge cells for header row 1
            sheet.merge_cells("B7:D7")  # Hóa đơn chứng từ mua

            # Header row styling
            header_font = Font(name="Times New Roman", size=11, bold=True)
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            header_fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

            # Apply header row 1
            for cell_ref, value in headers_row1.items():
                cell = sheet[cell_ref]
                cell.value = value
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill

            # ROW 8: Table sub-headers
            headers_row2 = {
                "A8": "",  # STT continues
                "B8": "Seri",
                "C8": "Số HĐ",
                "D8": "Ngày HĐ",
                "E8": "",  # Tên người bán continues
                "F8": "",  # Mã số thuế continues
                "G8": "",  # Mặt hàng continues
                "H8": "",  # Doanh số mua chưa thuế continues
                "I8": "",  # Thuế suất continues
                "J8": "",  # Thuế GTGT continues
                "K8": "",  # Ghi chú continues
            }

            # Apply header row 2
            for cell_ref, value in headers_row2.items():
                cell = sheet[cell_ref]
                cell.value = value
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill

            # Set row height for header rows
            sheet.row_dimensions[7].height = 30
            sheet.row_dimensions[8].height = 20

            # Prepare data for report
            # Map DataFrame columns to report columns
            column_map = {
                "document_number": "Số HĐ",
                "document_type": "Seri",  # Using document_type for Seri
                "date": "Ngày HĐ",
                "counterparty_name": (
                    "Tên người bán" if report_type == "MUA VÀO" else "Tên người mua"
                ),
                "counterparty_tax_number": "Mã số thuế",
                "description": "Mặt hàng",
                "amount_before_tax": (
                    "Doanh số mua\nchưa thuế"
                    if report_type == "MUA VÀO"
                    else "Doanh số bán\nchưa thuế"
                ),
                "tax_rate": "Thuế\nsuất",
                "tax_amount": "Thuế\nGTGT",
            }

            # Format the dataframe for the report
            report_df = df.copy()

            # Add STT (row numbers)
            report_df.insert(0, "STT", range(1, len(report_df) + 1))

            # Format dates
            if "date" in report_df.columns:
                # Convert to datetime if not already
                if not pd.api.types.is_datetime64_any_dtype(report_df["date"]):
                    report_df["date"] = pd.to_datetime(
                        report_df["date"], errors="coerce"
                    )
                # Format as DD/MM/YY
                report_df["date"] = report_df["date"].dt.strftime("%d/%m/%y")

            # Start row for data
            data_start_row = 9

            # Cell styles for data
            data_font = Font(name="Times New Roman", size=11)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Alignment styles
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")

            # Write data rows
            for index, row in report_df.iterrows():
                excel_row = data_start_row + index

                # Set row height
                sheet.row_dimensions[excel_row].height = 22

                # STT (A)
                cell = sheet.cell(row=excel_row, column=1)
                cell.value = row["STT"]
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Seri (B)
                cell = sheet.cell(row=excel_row, column=2)
                cell.value = row.get("document_type", "")
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Số HĐ (C)
                cell = sheet.cell(row=excel_row, column=3)
                cell.value = row.get("document_number", "")
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Ngày HĐ (D)
                cell = sheet.cell(row=excel_row, column=4)
                cell.value = row.get("date", "")
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Tên người bán/mua (E)
                cell = sheet.cell(row=excel_row, column=5)
                cell.value = row.get("counterparty_name", "")
                cell.font = data_font
                cell.alignment = left_align
                cell.border = border

                # Mã số thuế (F)
                cell = sheet.cell(row=excel_row, column=6)
                tax_number = row.get("counterparty_tax_number", "")
                if pd.notna(tax_number):
                    cell.value = str(tax_number)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Mặt hàng (G)
                cell = sheet.cell(row=excel_row, column=7)
                cell.value = row.get("description", "")
                cell.font = data_font
                cell.alignment = left_align
                cell.border = border

                # Doanh số mua/bán chưa thuế (H)
                cell = sheet.cell(row=excel_row, column=8)
                amount = row.get("amount_before_tax", 0)
                if pd.notna(amount):
                    cell.value = float(amount)
                    cell.number_format = "#,##0"
                cell.font = data_font
                cell.alignment = right_align
                cell.border = border

                # Thuế suất (I)
                cell = sheet.cell(row=excel_row, column=9)
                tax_rate = row.get("tax_rate", 0)
                if pd.notna(tax_rate):
                    # Check if already percentage or decimal
                    if isinstance(tax_rate, (int, float)):
                        if tax_rate > 1:  # Assume percentage
                            cell.value = tax_rate
                            cell.number_format = "0"
                        else:  # Assume decimal
                            cell.value = tax_rate * 100
                            cell.number_format = "0"
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

                # Thuế GTGT (J)
                cell = sheet.cell(row=excel_row, column=10)
                tax_amount = row.get("tax_amount", 0)
                if pd.notna(tax_amount):
                    cell.value = float(tax_amount)
                    cell.number_format = "#,##0"
                cell.font = data_font
                cell.alignment = right_align
                cell.border = border

                # Ghi chú (K)
                cell = sheet.cell(row=excel_row, column=11)
                cell.value = ""
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border

            # Total row
            total_row = data_start_row + len(report_df)

            # Format for total row
            total_font = Font(name="Times New Roman", size=11, bold=True)
            double_bottom_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="double"),
            )

            # Merge cells for "Tổng cộng"
            sheet.merge_cells(f"A{total_row}:G{total_row}")

            # Set "Tổng cộng" text
            cell = sheet.cell(row=total_row, column=1)
            cell.value = "Tổng cộng"
            cell.font = total_font
            cell.alignment = center_align
            cell.border = double_bottom_border

            # Set border for merged cells
            for col in range(2, 8):
                cell = sheet.cell(row=total_row, column=col)
                cell.border = double_bottom_border

            # Add sum formula for amount before tax (H)
            cell = sheet.cell(row=total_row, column=8)
            cell.value = f"=SUM(H{data_start_row}:H{total_row-1})"
            cell.font = total_font
            cell.alignment = right_align
            cell.border = double_bottom_border
            cell.number_format = "#,##0"

            # Empty cell for tax rate (I)
            cell = sheet.cell(row=total_row, column=9)
            cell.border = double_bottom_border
            cell.font = total_font

            # Add sum formula for tax amount (J)
            cell = sheet.cell(row=total_row, column=10)
            cell.value = f"=SUM(J{data_start_row}:J{total_row-1})"
            cell.font = total_font
            cell.alignment = right_align
            cell.border = double_bottom_border
            cell.number_format = "#,##0"

            # Empty cell for notes (K)
            cell = sheet.cell(row=total_row, column=11)
            cell.border = double_bottom_border
            cell.font = total_font

            # Add company signature section
            footer_start = total_row + 2

            # Date line on the right side
            sheet.merge_cells(f"I{footer_start}:K{footer_start}")
            cell = sheet.cell(row=footer_start, column=9)
            cell.value = f"Ngày {datetime.now().day} tháng {datetime.now().month} năm {datetime.now().year}"
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Position titles (Kế toán trưởng and Giám đốc)
            sig_row = footer_start + 1

            # Left side - Kế toán trưởng
            sheet.merge_cells(f"A{sig_row}:E{sig_row}")
            cell = sheet.cell(row=sig_row, column=1)
            cell.value = "Kế toán trưởng"
            cell.font = Font(name="Times New Roman", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Right side - Giám đốc
            sheet.merge_cells(f"I{sig_row}:K{sig_row}")
            cell = sheet.cell(row=sig_row, column=9)
            cell.value = "Giám đốc"
            cell.font = Font(name="Times New Roman", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add space for signatures
            sheet.row_dimensions[sig_row + 1].height = 20
            sheet.row_dimensions[sig_row + 2].height = 20
            sheet.row_dimensions[sig_row + 3].height = 20

            # Add name placeholders (if needed)
            name_row = sig_row + 4

            # Space for actual names
            # You can add names here if known, otherwise leave blank for manual entry
            # Example:
            # sheet.merge_cells(f'I{name_row}:K{name_row}')
            # cell = sheet.cell(row=name_row, column=9)
            # cell.value = "PHẠM VĂN CHƯƠNG"  # Example name
            # cell.font = Font(name="Times New Roman", size=11, bold=True)
            # cell.alignment = Alignment(horizontal="center", vertical="center")

            # Save the workbook
            wb.save(output_file)
            self.logger.info(f"Report saved as {output_file}")
            return True
        except Exception as e:
            self.logger.error(f"Error creating report: {str(e)}")
            return False
