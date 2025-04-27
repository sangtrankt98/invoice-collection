import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from utils.logger_setup import setup_logger


class InvoiceExcelGenerator:
    """
    A class to generate Excel files for invoice records (Vietnamese invoice listing).
    """

    def __init__(self, company_name="", tax_code="", address="", period="", year=None):
        """
        Initialize the generator with company information.

        Args:
            company_name (str): Name of the company
            tax_code (str): Tax code of the company
            address (str): Address of the company
            period (str): Period of the report (e.g., 'q4')
            year (int): Year of the report (defaults to current year if None)
        """
        self.logger = setup_logger()
        self.logger.info("Initializing InvoiceExcelGenerator")

        self.company_name = company_name
        self.tax_code = tax_code
        self.address = address
        self.period = period
        self.year = year if year else datetime.datetime.now().year

        # Initialize workbook and sheet
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "invoice_summarize"

        # Define styles
        self.header_font = Font(name="Times New Roman", bold=True, size=11)
        self.normal_font = Font(name="Times New Roman", size=10)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.left_alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        self.right_alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True
        )

        self.logger.debug(
            f"Generator initialized with company: {company_name}, period: {period}, year: {self.year}"
        )

    def create_header(self):
        """Create the header section of the template."""
        self.logger.info("Creating header section")
        try:
            # Title
            self.ws.merge_cells("A1:P1")
            self.ws["A1"] = "BẢNG KÊ HÓA ĐƠN CHỨNG TỪ HÀNG HÓA, DỊCH VỤ MUA VÀO"
            self.ws["A1"].font = self.header_font
            self.ws["A1"].alignment = self.center_alignment

            # Subtitle
            self.ws.merge_cells("A2:P2")
            self.ws["A2"] = "Kèm theo tờ khai thuế GTGT"
            self.ws["A2"].alignment = self.center_alignment

            self.ws.merge_cells("A3:P3")
            self.ws["A3"] = "(Dùng cho cơ sở kê khai khấu trừ thuế hàng tháng)"
            self.ws["A3"].alignment = self.center_alignment

            # Period
            self.ws.merge_cells("A4:P4")
            self.ws["A4"] = f"Kỳ {self.period} Năm {self.year}"
            self.ws["A4"].alignment = self.center_alignment

            # Company info
            self.ws.merge_cells("A5:K5")
            self.ws["A5"] = f"Tên cơ sở kinh doanh : {self.company_name}"
            self.ws["A5"].alignment = self.left_alignment

            self.ws.merge_cells("L5:M5")
            self.ws["L5"] = "Mã số thuế  :"
            self.ws["L5"].alignment = self.left_alignment

            self.ws.merge_cells("N5:P5")
            self.ws["N5"] = self.tax_code
            self.ws["N5"].alignment = self.left_alignment

            self.ws.merge_cells("A6:P6")
            self.ws["A6"] = f"Địa chỉ  : {self.address}"
            self.ws["A6"].alignment = self.left_alignment

            self.logger.debug("Header section created successfully")
        except Exception as e:
            self.logger.error(f"Error creating header section: {str(e)}")
            raise

    def create_table_headers(self):
        """Create the table headers."""
        self.logger.info("Creating table headers")
        try:
            headers = [
                "STT",
                "Chứng từ",
                "",
                "",
                "",
                "",
                "Hóa đơn chứng từ mua",
                "",
                "",
                "Tên người bán",
                "Mã sô thuế",
                "Mặt hàng",
                "Doanh số mua",
                "Thuế",
                "Thuế",
                "Ghi chú",
            ]

            # Merge cells for main headers
            self.ws.merge_cells("B7:F7")
            self.ws.merge_cells("G7:I7")

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=7, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border

            # Second row of headers
            sub_headers = [
                "",
                "Loại",
                "Số",
                "Ngày",
                "",
                "",
                "Seri",
                "Số HĐ",
                "Ngày HĐ",
                "",
                "",
                "",
                "chưa thuế",
                "suất",
                "GTGT",
                "",
            ]

            for col, header in enumerate(sub_headers, 1):
                cell = self.ws.cell(row=8, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border

            self.logger.debug("Table headers created successfully")
        except Exception as e:
            self.logger.error(f"Error creating table headers: {str(e)}")
            raise

    def add_data(self, data):
        """
        Add data to the template.

        Args:
            data (pandas.DataFrame): DataFrame with invoice data.
                Must contain columns matching the template structure.
        """
        self.logger.info(f"Adding data rows: {len(data)} rows")
        try:
            # Starting row for data
            row_idx = 9

            # Add each row of data
            for idx, row in data.iterrows():
                self.logger.debug(
                    f"Processing row {idx + 1}, invoice number: {row.get('invoice_number', 'N/A')}"
                )

                self.ws.cell(row=row_idx, column=1).value = idx + 1  # STT
                self.ws.cell(row=row_idx, column=2).value = row.get(
                    "document_type", ""
                )  # Loại
                self.ws.cell(row=row_idx, column=3).value = row.get(
                    "document_number", ""
                )  # Số
                self.ws.cell(row=row_idx, column=4).value = row.get(
                    "document_date", ""
                )  # Ngày
                # Column 5 and 6 are empty in the template
                self.ws.cell(row=row_idx, column=7).value = row.get(
                    "invoice_series", ""
                )  # Seri
                self.ws.cell(row=row_idx, column=8).value = row.get(
                    "invoice_number", ""
                )  # Số HĐ
                self.ws.cell(row=row_idx, column=9).value = row.get(
                    "invoice_date", ""
                )  # Ngày HĐ
                self.ws.cell(row=row_idx, column=10).value = row.get(
                    "seller_name", ""
                )  # Tên người bán
                self.ws.cell(row=row_idx, column=11).value = row.get(
                    "tax_code", ""
                )  # Mã số thuế
                self.ws.cell(row=row_idx, column=12).value = row.get(
                    "product_description", ""
                )  # Mặt hàng
                self.ws.cell(row=row_idx, column=13).value = row.get(
                    "amount_before_tax", 0
                )  # Doanh số chưa thuế
                self.ws.cell(row=row_idx, column=14).value = row.get(
                    "tax_rate", 0
                )  # Thuế suất
                self.ws.cell(row=row_idx, column=15).value = row.get(
                    "vat_amount", 0
                )  # Thuế GTGT
                self.ws.cell(row=row_idx, column=16).value = row.get(
                    "notes", ""
                )  # Ghi chú

                # Apply styles to all cells in the row
                for col in range(1, 17):
                    cell = self.ws.cell(row=row_idx, column=col)
                    cell.font = self.normal_font
                    cell.border = self.border

                    # Apply specific alignment based on column type
                    if col in [1, 14]:  # STT, tax rate
                        cell.alignment = self.center_alignment
                    elif col in [13, 15]:  # Amount columns
                        cell.alignment = self.right_alignment
                        # Format as number with comma separators
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0"
                    else:
                        cell.alignment = self.left_alignment

                row_idx += 1

            # Add total row
            self.logger.debug("Adding total row")
            total_row = row_idx
            self.ws.merge_cells(f"A{total_row}:E{total_row}")
            self.ws.cell(row=total_row, column=1).value = "TỔNG CỘNG"
            self.ws.cell(row=total_row, column=1).alignment = self.center_alignment
            self.ws.cell(row=total_row, column=1).font = self.header_font

            # Calculate totals
            try:
                total_amount = sum(data.get("amount_before_tax", 0))
                total_vat = sum(data.get("vat_amount", 0))
                self.logger.debug(
                    f"Calculated totals - Amount: {total_amount}, VAT: {total_vat}"
                )
            except Exception as calc_error:
                self.logger.warning(
                    f"Error calculating totals: {str(calc_error)}, using 0"
                )
                total_amount = 0
                total_vat = 0

            # Add totals to the row
            self.ws.cell(row=total_row, column=13).value = total_amount
            self.ws.cell(row=total_row, column=13).number_format = "#,##0"
            self.ws.cell(row=total_row, column=13).alignment = self.right_alignment
            self.ws.cell(row=total_row, column=13).font = self.header_font

            self.ws.cell(row=total_row, column=15).value = total_vat
            self.ws.cell(row=total_row, column=15).number_format = "#,##0"
            self.ws.cell(row=total_row, column=15).alignment = self.right_alignment
            self.ws.cell(row=total_row, column=15).font = self.header_font

            # Apply borders to total row
            for col in range(1, 17):
                self.ws.cell(row=total_row, column=col).border = self.border

            self.logger.info(
                f"Data added successfully: {len(data)} rows plus total row"
            )
        except Exception as e:
            self.logger.error(f"Error adding data: {str(e)}")
            raise

    def add_signature_section(self):
        """Add the signature section at the bottom of the template."""
        self.logger.info("Adding signature section")
        try:
            last_row = self.ws.max_row + 1

            # Date line
            self.ws.merge_cells(f"M{last_row}:P{last_row}")
            current_date = datetime.datetime.now()
            date_text = f"Ngày {current_date.day} tháng {current_date.month} năm {current_date.year}"
            self.ws.cell(row=last_row, column=13).value = date_text
            self.ws.cell(row=last_row, column=13).alignment = self.center_alignment

            # Signature titles
            last_row += 1
            self.ws.cell(row=last_row, column=2).value = "Người lập"
            self.ws.cell(row=last_row, column=2).alignment = self.center_alignment
            self.ws.cell(row=last_row, column=2).font = self.header_font

            self.ws.cell(row=last_row, column=8).value = "Kế toán trưởng"
            self.ws.cell(row=last_row, column=8).alignment = self.center_alignment
            self.ws.cell(row=last_row, column=8).font = self.header_font

            self.ws.cell(row=last_row, column=14).value = "Giám đốc"
            self.ws.cell(row=last_row, column=14).alignment = self.center_alignment
            self.ws.cell(row=last_row, column=14).font = self.header_font

            # Space for signatures (4 rows)
            last_row += 5

            # Signer name
            self.ws.cell(row=last_row, column=14).value = "PHẠM VĂN CHƯƠNG"
            self.ws.cell(row=last_row, column=14).alignment = self.center_alignment
            self.ws.cell(row=last_row, column=14).font = self.header_font

            self.logger.debug("Signature section added successfully")
        except Exception as e:
            self.logger.error(f"Error adding signature section: {str(e)}")
            raise

    def adjust_column_widths(self):
        """Adjust column widths to match template."""
        self.logger.info("Adjusting column widths")
        try:
            column_widths = {
                1: 4,  # STT
                2: 7,  # Loại
                3: 10,  # Số
                4: 10,  # Ngày
                5: 5,  # (empty)
                6: 5,  # (empty)
                7: 10,  # Seri
                8: 10,  # Số HĐ
                9: 10,  # Ngày HĐ
                10: 40,  # Tên người bán
                11: 15,  # Mã số thuế
                12: 30,  # Mặt hàng
                13: 15,  # Doanh số
                14: 8,  # Thuế suất
                15: 15,  # Thuế GTGT
                16: 10,  # Ghi chú
            }

            for col, width in column_widths.items():
                self.ws.column_dimensions[get_column_letter(col)].width = width

            self.logger.debug("Column widths adjusted successfully")
        except Exception as e:
            self.logger.error(f"Error adjusting column widths: {str(e)}")
            raise

    def generate_excel(self, data_df, output_path):
        """
        Generate the complete Excel file with the given data.

        Args:
            data_df (pandas.DataFrame): DataFrame with invoice data
            output_path (str): Path to save the Excel file

        Returns:
            str: Path to the generated file
        """
        self.logger.info(f"Generating Excel file at {output_path}")
        try:
            # Check if data is valid
            if data_df is None or data_df.empty:
                self.logger.warning("Data DataFrame is empty or None")
            else:
                self.logger.info(f"Data contains {len(data_df)} records")

            # Create the template
            self.create_header()
            self.create_table_headers()
            self.add_data(data_df)
            self.add_signature_section()
            self.adjust_column_widths()

            # Save the workbook
            self.wb.save(output_path)
            self.logger.info(f"Excel file successfully generated at {output_path}")
            return output_path
        except Exception as e:
            self.logger.error(f"Failed to generate Excel file: {str(e)}")
            raise
